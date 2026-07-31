import unicodedata
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE = unicodedata.normalize(
    "NFD", "/Volumes/cowork/EV/전기차 등록 현황/서울"
)
SRC = BASE + "/" + unicodedata.normalize("NFD", "15~21년 행정동별 연료별 자동차 등록현황.xlsx")
REF = BASE + "/" + unicodedata.normalize(
    "NFD", "서울시 자치구 읍면동별 연료별 자동차 등록현황(행정동)(22년12월).xlsx"
)
OUT = BASE + "/" + unicodedata.normalize(
    "NFD", "서울시 자치구 읍면동별 연료별 자동차 등록현황(행정동)(21년12월).xlsx"
)

USE_TYPES = ["관용", "자가용", "영업용"]
CAR_TYPES = ["승용", "승합", "화물", "특수"]

FUEL_ORDER = [
    "휘발유", "경유", "엘피지", "전기", "휘발유(유연)", "휘발유(무연)",
    "CNG", "LNG",
    "하이브리드(휘발유+전기)", "하이브리드(경유+전기)", "하이브리드(LPG+전기)", "하이브리드(CNG+전기)",
    "수소", "수소전기", "기타연료",
]


def load_ref_gu_order():
    df = pd.read_excel(REF, sheet_name=0, header=None)
    data = df.iloc[9:].dropna(subset=[3])
    return list(dict.fromkeys(data[0].dropna()))


def main():
    df = pd.read_excel(SRC, sheet_name="시군구별 행정동별 용도별 차종별 연료별 건수", header=1)
    d = df[df["연월"] == 202112].copy()

    d["행정동라벨"] = d["행정동사용본거지"].where(
        d["행정동사용본거지"].notna(), "기타"
    )
    # 기타(코드만 있고 동 이름 없는) 행은 구 내에서 하나로 합침 -> 코드값을 -1로 통일해 정렬상 맨 앞에 오게 함
    d["동정렬코드"] = d["행정동사용본거지코드"].where(
        d["행정동사용본거지"].notna(), -1
    )
    d["동정렬코드"] = pd.to_numeric(d["동정렬코드"], errors="coerce")

    grouped = (
        d.groupby(["사용본거지시군구", "행정동라벨", "동정렬코드", "연료", "용도", "차종종별"])["건수"]
        .sum()
        .reset_index()
    )
    lookup = {}
    for _, r in grouped.iterrows():
        key = (r["사용본거지시군구"], r["행정동라벨"], r["동정렬코드"], r["연료"])
        lookup.setdefault(key, {})[(r["용도"], r["차종종별"])] = int(r["건수"])

    gu_order = load_ref_gu_order()
    gu_rank = {g: i for i, g in enumerate(gu_order)}
    fuel_rank = {f: i for i, f in enumerate(FUEL_ORDER)}

    keys = list(lookup.keys())
    missing_fuel = {k[3] for k in keys if k[3] not in fuel_rank}
    if missing_fuel:
        raise ValueError(f"FUEL_ORDER에 없는 연료 발견: {missing_fuel}")
    missing_gu = {k[0] for k in keys if k[0] not in gu_rank}
    if missing_gu:
        raise ValueError(f"gu_order에 없는 시군구 발견: {missing_gu}")

    keys.sort(key=lambda k: (gu_rank[k[0]], k[2], fuel_rank[k[3]]))

    rows = []
    prev_gu = None
    prev_dong = None
    for gu, dong, _dong_code, fuel in keys:
        counts = lookup[(gu, dong, _dong_code, fuel)]
        vals = []
        for use in USE_TYPES:
            for car in CAR_TYPES:
                vals.append(counts.get((use, car), 0))
        total = sum(vals)
        if total == 0:
            continue

        show_gu = gu if gu != prev_gu else None
        # 구가 바뀌면 동도 새로 표시, 구가 같아도 동이 바뀌면 표시
        if gu != prev_gu or dong != prev_dong:
            show_dong = dong
        else:
            show_dong = None

        rows.append([show_gu, None, show_dong, fuel, total] + vals)
        prev_gu = gu
        prev_dong = dong

    # ---- 엑셀 작성 ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"

    ws.append([None] * 17)
    ws.append(["읍면동별 연료별 자동차 등록현황 (행정동)"] + [None] * 16)
    ws.append(["자동차관리정보시스템"] + [None] * 16)
    ws.append([None] * 17)
    row5 = [None] * 17
    row5[0] = "PROG_ID   :"
    row5[1] = "STA029Q21"
    row5[13] = "Page No.:"
    row5[15] = 1
    ws.append(row5)
    ws.append([None] * 17)
    row7 = [None] * 17
    row7[0] = "기준일자    :"
    row7[1] = "202112"
    row7[13] = "출력일시 :"
    ws.append(row7)

    header1 = [None] * 17
    header1[0] = "사용본거지 시군구"
    header1[2] = "읍면동 (행정동)"
    header1[3] = "연료"
    header1[4] = "계"
    header1[5] = "관용"
    header1[9] = "자가용"
    header1[13] = "영업용"
    ws.append(header1)

    header2 = [None] * 17
    for base_col in (5, 9, 13):
        for i, car in enumerate(CAR_TYPES):
            header2[base_col + i] = car
    ws.append(header2)

    for row in rows:
        ws.append(row)

    bold = Font(bold=True)
    ws["A2"].font = Font(bold=True, size=14)
    for cell in ["A8", "C8", "D8", "E8", "F8", "J8", "N8"]:
        ws[cell].font = bold
    for c in range(1, 18):
        ws.cell(row=9, column=c).font = bold

    for i in range(1, 18):
        ws.column_dimensions[get_column_letter(i)].width = 14

    wb.save(OUT)
    print("저장 완료:", OUT)
    print("총 데이터 행수:", len(rows))
    print("총 건수 합계:", sum(sum(r[5:]) for r in rows))


if __name__ == "__main__":
    main()
